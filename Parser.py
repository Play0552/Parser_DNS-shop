from time import sleep
import os
import time
import requests
from qrator_jsid import get_qrator_jsid
from bs4 import BeautifulSoup
import openpyxl


headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'ru,en;q=0.9',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.5845.2271 YaBrowser/23.9.0.2271 Yowser/2.5 Safari/537.36',
    'sec-ch-ua': '"Chromium";v="116", "Not)A;Brand";v="24", "YaBrowser";v="23"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
}


def get_page(url: str, cookies=None, params=None):
    """
    Получаем html страницу от сайта, при неудаче генерируем новый qrator_jsid
    """
    response = requests.get(url, cookies=cookies, params=params, headers=headers)
    if response.status_code != 200:
        qrator_jsid = get_qrator_jsid(3)
        cookies = {
            '_csrf': '3147e15b0fb56e58e4f6b3cdc47ddbfeb9b9c51f53babbac0840212086deeff9a%3A2%3A%7Bi%3A0%3Bs%3A5%3A%22_csrf%22%3Bi%3A1%3Bs%3A32%3A%22W0K1n5_Cv8QTGGmUMM39gG8itThhklO3%22%3B%7D',
            'qrator_jsid': qrator_jsid,
        }
        response = requests.get(url, cookies=cookies, params=params, headers=headers)
    sleep(2)  # чтобы не бомбить запросами
    return response, cookies


def get_product_param(data=None, cookies=None):
    """
    Получаем json из retail-rocket-product, при неудаче генерируем новый qrator_jsid
    """
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'ru,en;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Origin': 'https://www.dns-shop.ru',
        'Referer': 'https://www.dns-shop.ru/product/e3d62308868ded20/ugloslifovalnaa-masina-usm-finepower-ag90-125/no-referrer',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.5845.2271 YaBrowser/23.9.0.2271 Yowser/2.5 Safari/537.36',
        'X-CSRF-Token': 's8nUHS6TMDEFRqGtdCtJumCONkC71JoIbYdG40RoHqfZjKwoX6ZdXkE_6ZkNeg7dTeVfJu-e82Ak0w6rdV4vlQ==',
        'X-Requested-With': 'XMLHttpRequest',
        'content-type': 'application/x-www-form-urlencoded',
        'sec-ch-ua': '"Chromium";v="116", "Not)A;Brand";v="24", "YaBrowser";v="23"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }
    response = requests.post('https://www.dns-shop.ru/ajax-state/retail-rocket-product/', cookies=cookies, headers=headers, data=data)
    if response.status_code != 200:
        qrator_jsid = get_qrator_jsid(3)
        cookies = {
            '_csrf': '3147e15b0fb56e58e4f6b3cdc47ddbfeb9b9c51f53babbac0840212086deeff9a%3A2%3A%7Bi%3A0%3Bs%3A5%3A%22_csrf%22%3Bi%3A1%3Bs%3A32%3A%22W0K1n5_Cv8QTGGmUMM39gG8itThhklO3%22%3B%7D',
            'qrator_jsid': qrator_jsid,
        }
        response = requests.post('https://www.dns-shop.ru/ajax-state/retail-rocket-product/', cookies=cookies, headers=headers, data=data)
    sleep(2)  # чтобы не бомбить запросами
    return response, cookies

def get_url_from_page(soup):
    """
    Получаем все ссылки на товары со страницы
    """
    all_products_url_from_page = []
    all_url_from_page = soup.find_all('a', {'class': 'catalog-product__name'})
    for url in all_url_from_page:
        all_products_url_from_page.append(url.get('href'))
    return all_products_url_from_page


def take_all_products_url(cat_url: str, qrator_jsid=None, cookies=None):
    """
    Получаем все ссылки со всех страниц, а так же категорию
    """
    all_products_url = []

    if not qrator_jsid:
        qrator_jsid = get_qrator_jsid(3)
        cookies = {
            '_csrf': '3147e15b0fb56e58e4f6b3cdc47ddbfeb9b9c51f53babbac0840212086deeff9a%3A2%3A%7Bi%3A0%3Bs%3A5%3A%22_csrf%22%3Bi%3A1%3Bs%3A32%3A%22W0K1n5_Cv8QTGGmUMM39gG8itThhklO3%22%3B%7D',
            'qrator_jsid': qrator_jsid,
        }

    params = {
        'stock': 'now-out_of_stock',
    }

    response, cookies = get_page(cat_url, cookies=cookies, params=params)

    soup = BeautifulSoup(response.text, "lxml")
    category = soup.find('div', {'class': 'products-page__title'}).find('h1').text
    all_products_url.extend(get_url_from_page(soup))
    have_any_page = soup.find('a', {'class': 'pagination-widget__page-link_last'})
    if have_any_page:
        total_pages = int(soup.find('a', {'class': 'pagination-widget__page-link_last'}).get('href').split('=')[-1])
        if total_pages > 1:
            # Проходимся по всем страницам и записываем с них url товаров
            for page in range(2, total_pages+1):
                params = {
                    'stock': 'now-out_of_stock',
                    'p': f'{page}',
                }
                response, cookies = get_page(cat_url, cookies=cookies, params=params)
                soup = BeautifulSoup(response.text, "lxml")
                all_products_url.extend(get_url_from_page(soup))
    return all_products_url, cookies, category


def take_info_from_product_page(url: str, cookies=None):
    """
    Получаем всю информацию о продукте
    """
    link = f'https://www.dns-shop.ru{url}characteristics/'
    response, cookies = get_page(link, cookies=cookies)
    soup = BeautifulSoup(response.text, 'lxml')
    characteristics = dict(
        zip((_.text.strip() for _ in soup.find_all('div', {'class': 'product-characteristics__spec-title'})),
            (_.text.strip() for _ in soup.find_all('div', {'class': 'product-characteristics__spec-value'}))))
    product = soup.find('div', {'class': 'product-card'}).get('data-product-card').strip()
    script = soup.find_all('script')
    # вытаскиваем id который каждый раз генерируется для товара
    for string in script:
        if '"type":"retail-rocket-product"' in string.text:
            product_id = string.text.split('"type":"retail-rocket-product"')[1][10:19]
    # вытаскиваем ссылки на картинки и берём только thumb, можно менять на desktop или mobile, разница в размере
    for string in script:
        if 'window.initProductImagesSlider' in string.text:
            dict_list = eval(string.text.split(',"has3d"', 1)[0].split('"images":')[1])
            all_img_link = []
            for dct in dict_list:
                all_img_link.append(dct['thumb'].replace('\\', ''))

    # формируем data необходимую для retail-rocket-product
    data = f'data={{"type":"retail-rocket-product","containers":[{{"id":"{product_id}","data":{{"product":"{product}","requestUrl":"{url}"}}}}]}}'
    response, cookies = get_product_param(data=data, cookies=cookies)
    json_res = response.json()['data']['states'][0]['data']['data']
    price = json_res['price']
    main_image = json_res['pictureUrl']
    available = json_res['isAvailable']
    description = json_res['description']
    name = json_res['name']
    all_data = {
        'name': name,
        'price': price,
        'available': available,
        'link': link,
        'main_image': main_image,
        'all_img_link': all_img_link,
        'characteristics': characteristics,
        'description': description,
    }
    return all_data, cookies


def create_excel():
    """Создаёт excel файл с названием Parsing.xlsx"""
    book = openpyxl.Workbook()
    book.save('Parsing.xlsx')
    book.close()


def create_sheet(sheet):
    """Создаёт лист и записывает в него оглавления"""
    title_to_excel = ['Категория', 'Наименование', 'Цена', 'доступен или нет к продаже',
                      'Ссылка страницы с товаром', 'Ссылка на главное изображение',
                      'Ссылки на все изображения', 'Характеристики', 'Описание']
    wb = openpyxl.load_workbook('Parsing.xlsx')
    wb.create_sheet(sheet)
    ws = wb[sheet]
    ws.append(title_to_excel)
    wb.save('Parsing.xlsx')
    wb.close()


def excel_save(data: list, sheet: str):
    """Сохраняет данные в excel"""
    wb = openpyxl.load_workbook('Parsing.xlsx')
    ws = wb[sheet]
    ws.append(data)
    wb.save('Parsing.xlsx')
    wb.close()


def excel_del_sheet(sheet: str):
    """Удаляет лист из файла Parsing.xlsx"""
    wb = openpyxl.load_workbook('Parsing.xlsx')
    if sheet in wb.sheetnames:
        del wb[sheet]
    wb.save('Parsing.xlsx')
    wb.close()


def take_excel():
    """
    Создаём excel файл если его нет или если он создан более 24ч назад и заполняем,
    иначе возвращаем существующий файл
    """
    day = 86400
    if (not os.path.exists('Parsing.xlsx') or
            ((time.time() - os.path.getmtime('Parsing.xlsx')) > day)):
        create_excel()
        cats_url = ['https://www.dns-shop.ru/catalog/recipe/877f4d35bf74c8b4/derzateli-dla-zubnyh-setok/',
                    'https://www.dns-shop.ru/catalog/2a8f00c7c701409e/derzhateli-dlya-videokart/',
                    'https://www.dns-shop.ru/catalog/15efef292eb04e77/vstraivaemye-kofemashiny/',
                    'https://www.dns-shop.ru/catalog/17a9d40a16404e77/nastolnye-chasy/',
                    'https://www.dns-shop.ru/catalog/dc4dfa0e5f7e7fd7/metalloiskateli/']

        qrator_jsid = '1696175207.460.y3aLqV96XSTGwbmp-9lt9l0s81mdoc12nsug9a7bebcblgs0j'
        cookies = {
            '_csrf': '3147e15b0fb56e58e4f6b3cdc47ddbfeb9b9c51f53babbac0840212086deeff9a%3A2%3A%7Bi%3A0%3Bs%3A5%3A%22_csrf%22%3Bi%3A1%3Bs%3A32%3A%22W0K1n5_Cv8QTGGmUMM39gG8itThhklO3%22%3B%7D',
            'qrator_jsid': qrator_jsid,
            'city_path': 'barnaul',
        }
        total_cats = len(cats_url)
        counter_cats = total_cats
        for cat_url in cats_url:
            print(f'Осталось категорий: {counter_cats} из {total_cats}')
            counter_cats -= 1
            all_products_url, cookies, category = take_all_products_url(cat_url, qrator_jsid, cookies=cookies)
            create_sheet(category)
            total_products = len(all_products_url)
            counter_prod = total_products
            for url in all_products_url:
                try:
                    all_data, cookies = take_info_from_product_page(url, cookies=cookies)
                except Exception as e:
                    print(e)
                else:
                    to_excel = [category, all_data['name'], all_data['price'], str(all_data['available']),
                                all_data['link'], all_data['main_image'], str(all_data['all_img_link']),
                                str(all_data['characteristics']), all_data['description']]
                    excel_save(to_excel, category)
                print(f'Осталось: {counter_prod} из {total_products} в категории: {category}')
                counter_prod -= 1
        #Удалим начальную страницу в экселе
        excel_del_sheet('Sheet')
    return


if __name__ == '__main__':
    take_excel()
